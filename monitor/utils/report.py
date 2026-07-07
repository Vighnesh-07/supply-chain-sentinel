"""
Supply Chain Sentinel — Report Generator
==========================================
Produces both rich terminal summaries and professionally formatted
.xlsx reports with severity-coded coloring, freeze panes, and
summary statistics.

Supports multi-ecosystem grouping and multi-database threat source attribution.
"""

import os
from datetime import datetime
from typing import Dict, List

import pandas as pd

from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.columns import Columns
from rich.text import Text

from .config import SEVERITY_STYLES, SEVERITY_ICONS

console = Console()


# ═══════════════════════════════════════════════════════════════
# TERMINAL DISPLAY
# ═══════════════════════════════════════════════════════════════

def _severity_style(severity: str) -> str:
    """Map severity to Rich console style."""
    return SEVERITY_STYLES.get(severity, "white")


def _build_severity_bar(count: int, total: int, label: str, style: str) -> str:
    """Build a visual bar for severity distribution."""
    if total == 0:
        pct = 0
    else:
        pct = int((count / total) * 20)
    bar = "=" * pct + " " * (20 - pct)
    pct_val = (count / total * 100) if total > 0 else 0
    return f"[{style}]{label:>10}[/] [{style}]\\[{bar}][/] [bold]{count}[/] ({pct_val:.0f}%)"


def _build_ecosystem_summary(audit_data: List[Dict]) -> str:
    """Build a per-ecosystem package count summary string."""
    eco_counts: Dict[str, int] = {}
    for item in audit_data:
        eco = item.get("ecosystem", "Unknown")
        eco_counts[eco] = eco_counts.get(eco, 0) + 1

    parts = []
    for eco, count in sorted(eco_counts.items(), key=lambda x: -x[1]):
        parts.append(f"[bold]{eco}[/]: {count}")
    return "  |  ".join(parts)


def display_terminal_summary(audit_data: List[Dict]) -> None:
    """
    Display audit results as a comprehensive security dashboard in the terminal.

    Args:
        audit_data: List of audit result dicts (now includes 'ecosystem' and 'threat_sources').
    """
    console.print("\n")

    # ════════════════════════════════════════════════════════════
    # SECTION 1: Full Audit Results Table
    # ════════════════════════════════════════════════════════════
    table = Table(
        show_header=True,
        header_style="bold white on dark_blue",
        border_style="bright_blue",
        show_lines=True,
        padding=(0, 1),
        min_width=120,
        expand=False,
    )

    table.add_column("#", style="dim", width=4, justify="right")
    table.add_column("Package", style="bold white", min_width=22)
    table.add_column("Ecosystem", style="bold magenta", justify="center", min_width=14)
    table.add_column("Installed", style="cyan", justify="center", min_width=10)
    table.add_column("Latest", justify="center", min_width=10)
    table.add_column("Malicious", justify="center", min_width=11)
    table.add_column("Severity", justify="center", min_width=14)
    table.add_column("CVSS", justify="center", min_width=6)
    table.add_column("Vulns", justify="center", min_width=6)
    table.add_column("IOC Risk", justify="center", min_width=18)
    table.add_column("CVE IDs", style="dim", min_width=18)

    # Sort by severity (most severe first)
    severity_order = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "NONE": 4}
    sorted_data = sorted(
        audit_data,
        key=lambda x: severity_order.get(x.get("severity", "NONE"), 5),
    )

    for idx, item in enumerate(sorted_data, 1):
        severity = item.get("severity", "NONE")
        sev_icon = SEVERITY_ICONS.get(severity, "")
        severity_display = f"[{_severity_style(severity)}]{sev_icon} {severity}[/]"

        current = item.get("current_version", "unknown")
        latest = item.get("latest_version", "N/A") or "N/A"
        ecosystem = item.get("ecosystem", "Unknown")

        # Indicate version upgrade available
        if latest != "N/A" and latest != current:
            latest_display = f"[bold green]>> {latest}[/]"
        else:
            latest_display = f"[dim]{latest}[/]"

        # Format CVE list
        cves = item.get("cve_ids", [])
        if len(cves) > 2:
            cve_display = ", ".join(cves[:2]) + f"\n(+{len(cves) - 2} more)"
        elif cves:
            cve_display = ", ".join(cves)
        else:
            cve_display = "[dim]--[/]"

        # Malicious Flag with source attribution
        is_mal = item.get("is_malicious", False)
        threat_sources = item.get("threat_sources", [])
        if is_mal:
            mal_display = "[bold white on red] MALICIOUS [/]"
        else:
            mal_display = "[dim green]Clean[/]"

        # CVSS
        cvss = item.get("cvss_score", 0.0)
        if cvss >= 9.0:
            cvss_display = f"[bold red]{cvss:.1f}[/]"
        elif cvss >= 7.0:
            cvss_display = f"[red]{cvss:.1f}[/]"
        elif cvss >= 4.0:
            cvss_display = f"[yellow]{cvss:.1f}[/]"
        elif cvss > 0:
            cvss_display = f"[blue]{cvss:.1f}[/]"
        else:
            cvss_display = "[dim]0.0[/]"

        # Vuln count
        vuln_count = item.get("vuln_count", 0)
        vuln_display = f"[bold red]{vuln_count}[/]" if vuln_count > 0 else "[dim green]0[/]"

        # IOC Risk summary
        risk_summary = item.get("risk_summary", "")
        if risk_summary and risk_summary != "No IOCs found":
            ioc_display = f"[bold yellow]{risk_summary}[/]"
        else:
            ioc_display = "[dim green]Clean[/]"

        table.add_row(
            str(idx),
            item.get("package", ""),
            ecosystem,
            current,
            latest_display,
            mal_display,
            severity_display,
            cvss_display,
            vuln_display,
            ioc_display,
            cve_display,
        )

    console.print(Panel(
        table,
        title="[bold bright_cyan]COMPLETE AUDIT RESULTS[/]",
        border_style="bright_blue",
        padding=(1, 1),
    ))

    # ════════════════════════════════════════════════════════════
    # SECTION 2: Security Scorecard Dashboard
    # ════════════════════════════════════════════════════════════
    total = len(audit_data)
    critical = sum(1 for d in audit_data if d.get("severity") == "CRITICAL")
    high = sum(1 for d in audit_data if d.get("severity") == "HIGH")
    medium = sum(1 for d in audit_data if d.get("severity") == "MEDIUM")
    low = sum(1 for d in audit_data if d.get("severity") == "LOW")
    clean = sum(1 for d in audit_data if d.get("severity") == "NONE")
    malicious = sum(1 for d in audit_data if d.get("is_malicious"))
    outdated = sum(
        1 for d in audit_data
        if d.get("latest_version") and d["latest_version"] != "N/A"
        and d["latest_version"] != d.get("current_version")
    )
    vulnerable = total - clean
    total_cves = sum(len(d.get("cve_ids", [])) for d in audit_data)

    # Risk Score calculation (simple weighted)
    risk_score = (critical * 40) + (high * 20) + (medium * 5) + (low * 1) + (malicious * 100)
    if risk_score == 0:
        risk_grade = "A+"
        risk_color = "bold green"
    elif risk_score <= 20:
        risk_grade = "A"
        risk_color = "bold green"
    elif risk_score <= 50:
        risk_grade = "B"
        risk_color = "bold yellow"
    elif risk_score <= 100:
        risk_grade = "C"
        risk_color = "bold yellow"
    elif risk_score <= 200:
        risk_grade = "D"
        risk_color = "bold red"
    else:
        risk_grade = "F"
        risk_color = "bold white on red"

    # Determine border color
    if malicious > 0 or critical > 0:
        card_border = "bold red"
    elif high > 0:
        card_border = "yellow"
    else:
        card_border = "green"

    # Build the scorecard in a side-by-side grid
    left_grid = Table.grid(padding=(0, 1))
    left_grid.add_column("Key", style="bold cyan", no_wrap=True)
    left_grid.add_column("Val", style="bold white", no_wrap=True)
    left_grid.add_row("RISK GRADE", f"[{risk_color}]{risk_grade}  (Score: {risk_score})[/]")
    left_grid.add_row("Packages Scanned", str(total))
    left_grid.add_row("Malicious Packages", f"[bold red]{malicious}[/]" if malicious > 0 else "[bold green]0[/]")
    left_grid.add_row("Vulnerable Packages", f"[bold yellow]{vulnerable}[/]" if vulnerable > 0 else "[bold green]0[/]")
    left_grid.add_row("Outdated Packages", f"[bold yellow]{outdated}[/]" if outdated > 0 else "[bold green]0[/]")
    left_grid.add_row("Total CVEs Found", str(total_cves))
    left_grid.add_row("", "")
    if malicious > 0:
        left_grid.add_row("[bold white on red]STATUS[/]", "[bold white on red] !! MALICIOUS DETECTED !! [/]")
    else:
        left_grid.add_row("STATUS", "[bold green][OK] No malicious packages detected[/]")

    right_grid = Table.grid(padding=(0, 1))
    right_grid.add_column("Label")
    right_grid.add_row(" [bold underline cyan]Severity Distribution:[/]")
    right_grid.add_row(f"  {_build_severity_bar(critical, total, 'CRITICAL', 'bold red')}")
    right_grid.add_row(f"  {_build_severity_bar(high, total, 'HIGH', 'red')}")
    right_grid.add_row(f"  {_build_severity_bar(medium, total, 'MEDIUM', 'yellow')}")
    right_grid.add_row(f"  {_build_severity_bar(low, total, 'LOW', 'blue')}")
    right_grid.add_row(f"  {_build_severity_bar(clean, total, 'CLEAN', 'green')}")
    right_grid.add_row("")
    right_grid.add_row(" [bold underline cyan]Ecosystem Summary:[/]")
    right_grid.add_row(f"  {_build_ecosystem_summary(audit_data)}")

    scorecard_grid = Table.grid(expand=True, padding=(0, 4))
    scorecard_grid.add_column("Left", ratio=1, vertical="top")
    scorecard_grid.add_column("Right", ratio=1, vertical="top")
    scorecard_grid.add_row(left_grid, right_grid)

    console.print(Panel(
        scorecard_grid,
        title=f"[bold bright_cyan]SECURITY SCORECARD[/]",
        border_style=card_border,
        padding=(1, 2),
    ))

    # ════════════════════════════════════════════════════════════
    # SECTION 3: Threat Intelligence Source Summary
    # ════════════════════════════════════════════════════════════
    flagged_items = [d for d in audit_data if d.get("is_malicious") or d.get("threat_sources")]
    if flagged_items:
        threat_table = Table(
            show_header=True,
            header_style="bold white on red",
            border_style="red",
            padding=(0, 1),
            show_lines=True,
            expand=False,
        )
        threat_table.add_column("#", style="bold red", width=4, justify="right")
        threat_table.add_column("Package", style="bold white", min_width=20)
        threat_table.add_column("Ecosystem", style="bold magenta", min_width=14)
        threat_table.add_column("Status", style="bold red", min_width=12)
        threat_table.add_column("Threat Database / Source", min_width=35)

        for idx, item in enumerate(flagged_items, 1):
            sources = item.get("threat_sources", [])
            source_str = ", ".join(sources) if sources else "Unknown"
            status = "[bold white on red] MALICIOUS [/]" if item.get("is_malicious") else "[bold yellow]FLAGGED[/]"
            threat_table.add_row(
                str(idx),
                item.get("package", ""),
                item.get("ecosystem", "Unknown"),
                status,
                f"[red]{source_str}[/]",
            )

        console.print(Panel(
            threat_table,
            title=f"[bold white on red] THREAT INTELLIGENCE REPORT  |  {len(flagged_items)} FLAGGED [/]",
            border_style="bold red",
            padding=(1, 1),
        ))

    # ════════════════════════════════════════════════════════════
    # SECTION 4: Vulnerable Packages Detail (only if vulns exist)
    # ════════════════════════════════════════════════════════════
    vulnerable_items = [d for d in audit_data if d.get("vuln_count", 0) > 0]
    if vulnerable_items:
        console.print(f"\n  [bold bright_cyan]{'='*60}[/]")
        console.print(f"  [bold bright_cyan]  VULNERABLE PACKAGE DETAILS[/]")
        console.print(f"  [bold bright_cyan]{'='*60}[/]")

        for item in vulnerable_items:
            name = item["package"]
            severity = item.get("severity", "NONE")
            cvss = item.get("cvss_score", 0.0)
            cves = item.get("cve_ids", [])
            vuln_count = item.get("vuln_count", 0)
            ecosystem = item.get("ecosystem", "Unknown")
            sev_style = _severity_style(severity)
            sev_icon = SEVERITY_ICONS.get(severity, "")

            detail_lines = []
            detail_lines.append(f"  [bold]Package:[/]    {name}")
            detail_lines.append(f"  [bold]Ecosystem:[/]  {ecosystem}")
            detail_lines.append(f"  [bold]Version:[/]    {item.get('current_version', '?')}")
            detail_lines.append(f"  [bold]Severity:[/]   [{sev_style}]{sev_icon} {severity}[/]")
            detail_lines.append(f"  [bold]CVSS Score:[/] [{sev_style}]{cvss:.1f} / 10.0[/]")
            detail_lines.append(f"  [bold]Vuln Count:[/] [bold red]{vuln_count}[/]")

            if cves:
                detail_lines.append(f"  [bold]CVE IDs:[/]")
                for cve in cves:
                    detail_lines.append(f"    - [dim]{cve}[/]")

            if item.get("latest_version") and item["latest_version"] != "N/A":
                detail_lines.append(f"  [bold]Fix:[/]        Upgrade to [bold green]{item['latest_version']}[/]")

            panel_border = "red" if severity in ("CRITICAL", "HIGH") else "yellow"
            console.print(Panel(
                "\n".join(detail_lines),
                title=f"[{sev_style}]{sev_icon} {name}[/]",
                border_style=panel_border,
                padding=(0, 1),
            ))

    # ════════════════════════════════════════════════════════════
    # SECTION 5: Per-Package IOC Detailed Breakdown
    # ════════════════════════════════════════════════════════════
    ioc_items = [d for d in audit_data if d.get("risk_summary") and d["risk_summary"] != "No IOCs found"]
    if ioc_items:
        from rich.tree import Tree
        console.print(f"\n  [bold bright_magenta]{'='*60}[/]")
        console.print(f"  [bold bright_magenta]  PER-PACKAGE IOC BREAKDOWN[/]")
        console.print(f"  [bold bright_magenta]{'='*60}[/]")

        for item in ioc_items:
            name = item["package"]
            eco = item.get("ecosystem", "Unknown")
            
            tree = Tree(f"[bold yellow][!!] {name} ({eco})[/] - [yellow]{item.get('risk_summary', '')}[/]")
            
            def add_ioc_tree(title, items, style):
                if items:
                    branch = tree.add(f"[bold]{title}[/]")
                    for idx, val in enumerate(items[:5]):
                        if isinstance(val, dict):
                            cat = val.get("type", "Unknown")
                            desc = val.get("description", "")
                            branch.add(f"[{style}]{cat}[/]: {desc}")
                        else:
                            branch.add(f"[{style}]{val}[/]")
                    if len(items) > 5:
                        branch.add(f"[dim]...and {len(items)-5} more[/]")

            add_ioc_tree("AST Findings", item.get("ioc_ast", []), "cyan")
            add_ioc_tree("Crypto Wallets", item.get("ioc_wallets", []), "magenta")
            add_ioc_tree("High-Entropy Secrets", item.get("ioc_entropy_secrets", []), "red")
            add_ioc_tree("Suspicious URLs", item.get("ioc_urls", []), "blue")
            add_ioc_tree("Suspicious IPs", item.get("ioc_ips", []), "yellow")
            add_ioc_tree("Base64 Strings", item.get("ioc_base64", []), "dim")
            add_ioc_tree("Cloud Assets/Webhooks", item.get("ioc_sens", []), "green")

            console.print(Panel(
                tree,
                border_style="yellow",
                padding=(1, 2),
            ))

# ═══════════════════════════════════════════════════════════════
# EXCEL REPORT GENERATION
# ═══════════════════════════════════════════════════════════════

def generate_excel_report(audit_data: List[Dict], output_dir: str) -> str:
    """
    Generate a professionally formatted Excel (.xlsx) report.

    The report includes:
      - Color-coded severity column (CRITICAL=red, HIGH=orange, etc.)
      - Ecosystem column showing package origin
      - Threat Sources column showing which databases flagged each package
      - Styled header row and title banner
      - Summary statistics row
      - Frozen header panes for scrollability

    Args:
        audit_data: List of audit result dicts.
        output_dir: Directory to save the report.

    Returns:
        Absolute path to the generated .xlsx file.
    """
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"supply_chain_audit_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    # ── Build DataFrame ──
    df = pd.DataFrame([
        {
            "Component/Package": item["package"],
            "Ecosystem": item.get("ecosystem", "Unknown"),
            "Current Version": item["current_version"],
            "Latest Version Available": item.get("latest_version", "N/A") or "N/A",
            "Malicious Flag": item.get("is_malicious", False),
            "Threat Sources": ", ".join(item.get("threat_sources", [])) if item.get("threat_sources") else "--",
            "Severity Level": item.get("severity", "NONE"),
            "CVSS Score": item.get("cvss_score", 0.0),
            "CVE Count": item.get("vuln_count", 0),
            "CVE IDs": ", ".join(item.get("cve_ids", [])),
            "Flagged For": item.get("risk_summary", "Clean") if item.get("risk_summary") else "Clean",
            "Crypto Wallets": ", ".join(item.get("ioc_wallets", [])) if item.get("ioc_wallets") else "--",
            "High-Entropy Secrets": ", ".join(item.get("ioc_entropy_secrets", [])) if item.get("ioc_entropy_secrets") else "--",
            "AST Findings": ", ".join([a.get("type", str(a)) if isinstance(a, dict) else str(a) for a in item.get("ioc_ast", [])]) if item.get("ioc_ast") else "--",
            "Suspicious URLs": ", ".join(item.get("ioc_urls", [])) if item.get("ioc_urls") else "--",
            "Suspicious IPs": ", ".join(item.get("ioc_ips", [])) if item.get("ioc_ips") else "--",
            "Base64 Strings": ", ".join(item.get("ioc_base64", [])) if item.get("ioc_base64") else "--",
            "Hex Strings": ", ".join(item.get("ioc_hex", [])) if item.get("ioc_hex") else "--",
            "Firebase Findings": ", ".join(item.get("ioc_firebase", [])) if item.get("ioc_firebase") else "--",
            "S3 Buckets": ", ".join(item.get("ioc_s3", [])) if item.get("ioc_s3") else "--",
            "Cloud Assets / Webhooks": ", ".join(item.get("ioc_sens", [])) if item.get("ioc_sens") else "--",
            "AbuseIPDB Scores": ", ".join(item.get("abuseipdb_scores", [])) if item.get("abuseipdb_scores") else "--",
        }
        for item in audit_data
    ])

    # Sort by severity
    severity_order = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "NONE": 4}
    df["_sort"] = df["Severity Level"].map(severity_order)
    df = df.sort_values("_sort").drop("_sort", axis=1).reset_index(drop=True)

    # ── Write with openpyxl formatting ──
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Audit Results", index=False, startrow=2)

        workbook = writer.book
        worksheet = writer.sheets["Audit Results"]

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # ── Title Banner ──
        col_count = len(df.columns)
        last_col_letter = chr(ord("A") + col_count - 1)
        worksheet.merge_cells(f"A1:{last_col_letter}1")
        title_cell = worksheet["A1"]
        title_cell.value = (
            f"Supply Chain Security Audit Report (Multi-Ecosystem)  --  "
            f"{datetime.now().strftime('%Y-%m-%d %H:%M')}"
        )
        title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(
            start_color="1B2838", end_color="1B2838", fill_type="solid"
        )
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.row_dimensions[1].height = 38

        # ── Subtitle Row ──
        worksheet.merge_cells(f"A2:{last_col_letter}2")
        subtitle_cell = worksheet["A2"]
        subtitle_cell.value = "Generated by Supply Chain Sentinel  |  Multi-Ecosystem SBOM + Dependency Audit + Runtime Scan  |  Threat Intel: OSV + GitHub Advisory + npm Audit + Local Blocklist"
        subtitle_cell.font = Font(name="Calibri", size=9, italic=True, color="8899AA")
        subtitle_cell.fill = PatternFill(
            start_color="1B2838", end_color="1B2838", fill_type="solid"
        )
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.row_dimensions[2].height = 20

        # ── Header Row Styling (row 3) ──
        header_fill = PatternFill(
            start_color="2D4059", end_color="2D4059", fill_type="solid"
        )
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin", color="444444"),
            right=Side(style="thin", color="444444"),
            top=Side(style="thin", color="444444"),
            bottom=Side(style="thin", color="444444"),
        )

        for col_idx in range(1, col_count + 1):
            cell = worksheet.cell(row=3, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        worksheet.row_dimensions[3].height = 30

        # ── Severity Color Map ──
        severity_fills = {
            "CRITICAL": PatternFill(
                start_color="FF1744", end_color="FF1744", fill_type="solid"
            ),
            "HIGH": PatternFill(
                start_color="FF6D00", end_color="FF6D00", fill_type="solid"
            ),
            "MEDIUM": PatternFill(
                start_color="FFD600", end_color="FFD600", fill_type="solid"
            ),
            "LOW": PatternFill(
                start_color="448AFF", end_color="448AFF", fill_type="solid"
            ),
            "NONE": PatternFill(
                start_color="00E676", end_color="00E676", fill_type="solid"
            ),
        }
        severity_fonts = {
            "CRITICAL": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "HIGH": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "MEDIUM": Font(name="Calibri", size=10, bold=True, color="000000"),
            "LOW": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "NONE": Font(name="Calibri", size=10, bold=True, color="000000"),
        }

        # Ecosystem color map
        ecosystem_fills = {
            "PyPI": PatternFill(start_color="3776AB", end_color="3776AB", fill_type="solid"),
            "npm": PatternFill(start_color="CB3837", end_color="CB3837", fill_type="solid"),
            "Go": PatternFill(start_color="00ADD8", end_color="00ADD8", fill_type="solid"),
            "RubyGems": PatternFill(start_color="CC342D", end_color="CC342D", fill_type="solid"),
            "Maven": PatternFill(start_color="C71A36", end_color="C71A36", fill_type="solid"),
            "crates.io": PatternFill(start_color="DEA584", end_color="DEA584", fill_type="solid"),
            "Packagist": PatternFill(start_color="777BB4", end_color="777BB4", fill_type="solid"),
            "NuGet": PatternFill(start_color="004880", end_color="004880", fill_type="solid"),
        }
        ecosystem_font_white = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        ecosystem_font_dark = Font(name="Calibri", size=10, bold=True, color="000000")

        # Alternating row fill for readability
        row_fill_even = PatternFill(
            start_color="F5F7FA", end_color="F5F7FA", fill_type="solid"
        )
        row_fill_odd = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )

        severity_col_idx = df.columns.get_loc("Severity Level") + 1  # 1-indexed
        ecosystem_col_idx = df.columns.get_loc("Ecosystem") + 1
        data_start_row = 4  # Row 1=title, 2=subtitle, 3=header, 4=first data

        for row_offset in range(len(df)):
            row_idx = data_start_row + row_offset
            row_fill = row_fill_even if row_offset % 2 == 0 else row_fill_odd

            for col_idx in range(1, col_count + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Calibri", size=10)
                cell.fill = row_fill

            # Package name left-aligned and bold
            pkg_cell = worksheet.cell(row=row_idx, column=1)
            pkg_cell.alignment = Alignment(horizontal="left", vertical="center")
            pkg_cell.font = Font(name="Calibri", size=10, bold=True)

            # Apply ecosystem color coding
            eco_cell = worksheet.cell(row=row_idx, column=ecosystem_col_idx)
            eco_val = eco_cell.value
            if eco_val in ecosystem_fills:
                eco_cell.fill = ecosystem_fills[eco_val]
                eco_cell.font = ecosystem_font_dark if eco_val == "crates.io" else ecosystem_font_white

            # Apply severity color coding
            sev_cell = worksheet.cell(row=row_idx, column=severity_col_idx)
            severity_val = sev_cell.value
            if severity_val in severity_fills:
                sev_cell.fill = severity_fills[severity_val]
                sev_cell.font = severity_fonts[severity_val]

            # Apply malicious flag styling
            mal_col_idx = df.columns.get_loc("Malicious Flag") + 1
            mal_cell = worksheet.cell(row=row_idx, column=mal_col_idx)
            if mal_cell.value:
                mal_cell.fill = PatternFill(start_color="B71C1C", end_color="B71C1C", fill_type="solid")
                mal_cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
                mal_cell.value = "!! MALICIOUS"
            else:
                mal_cell.value = "CLEAN"

        # ── Column Widths ──
        column_widths = {
            "A": 30,  # Package name
            "B": 20,  # Ecosystem
            "C": 16,  # Current Version
            "D": 22,  # Latest Version
            "E": 16,  # Malicious Flag
            "F": 35,  # Threat Sources
            "G": 16,  # Severity
            "H": 12,  # CVSS Score
            "I": 12,  # CVE Count
            "J": 50,  # CVE IDs
            "K": 45,  # Suspicious URLs
            "L": 25,  # Suspicious IPs
            "M": 40,  # Base64 Strings
            "N": 40,  # Hex Strings
            "O": 35,  # Firebase Findings
            "P": 35,  # S3 Buckets
            "Q": 40,  # Cloud Assets / Webhooks
            "R": 30,  # AbuseIPDB Scores
        }
        for col_letter, width in column_widths.items():
            if col_letter in worksheet.column_dimensions:
                worksheet.column_dimensions[col_letter].width = width

        # ── Summary Footer ──
        summary_row = data_start_row + len(df) + 1
        worksheet.merge_cells(f"A{summary_row}:{last_col_letter}{summary_row}")
        summary_cell = worksheet.cell(row=summary_row, column=1)

        critical = sum(1 for _, r in df.iterrows() if r["Severity Level"] == "CRITICAL")
        high = sum(1 for _, r in df.iterrows() if r["Severity Level"] == "HIGH")
        medium = sum(1 for _, r in df.iterrows() if r["Severity Level"] == "MEDIUM")
        low = sum(1 for _, r in df.iterrows() if r["Severity Level"] == "LOW")
        clean = sum(1 for _, r in df.iterrows() if r["Severity Level"] == "NONE")

        # Count unique ecosystems
        eco_list = df["Ecosystem"].unique().tolist()
        eco_str = ", ".join(eco_list)

        summary_cell.value = (
            f"Total: {len(df)} packages across {len(eco_list)} ecosystems ({eco_str})  |  "
            f"Critical: {critical}  |  High: {high}  |  "
            f"Medium: {medium}  |  Low: {low}  |  Clean: {clean}  |  "
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        )
        summary_cell.font = Font(name="Calibri", size=10, italic=True, color="666666")
        summary_cell.alignment = Alignment(horizontal="center")
        summary_cell.fill = PatternFill(
            start_color="EAEEF3", end_color="EAEEF3", fill_type="solid"
        )

        # ── Freeze Panes (keep title + header visible) ──
        worksheet.freeze_panes = "A4"

    console.print(f"\n  [bold green][OK] Excel report saved:[/] {os.path.abspath(filepath)}")
    return os.path.abspath(filepath)
